import datetime
from output_record import OutputRecord
from lserm_row import LsermRow

import openpyxl.worksheet.worksheet
import openpyxl.styles
import openpyxl.worksheet.dimensions

class ExcelContext:
    def __init__(self, output_path):
        self.wb = None
        self.output_path = output_path

    def __enter__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        return self.wb

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(self.output_path)
        self.wb.close()

    @classmethod
    def write(cls, path:str, records:list[OutputRecord]):
        with cls(path) as wb:
            for i in records:
                o=OutputExcel.convert(i, wb)
                o.write()


class OutputExcel(OutputRecord):
    BOLD_FONT = openpyxl.styles.Font("Calibri", bold=True)
    BORDER_SIDE = openpyxl.styles.Side(style="thin")
    HEADER_LEFT = openpyxl.styles.Border(top=BORDER_SIDE, bottom=BORDER_SIDE, left=BORDER_SIDE)
    HEADER_RIGHT = openpyxl.styles.Border(top=BORDER_SIDE, bottom=BORDER_SIDE, right=BORDER_SIDE)
    ROW_LEFT = openpyxl.styles.Border(left=BORDER_SIDE)
    ROW_RIGHT = openpyxl.styles.Border(right=BORDER_SIDE)
    HIGHLIGHT = openpyxl.styles.PatternFill("solid", bgColor="fffff2cc", fgColor="fffff2cc")
    DATE_FORMAT = 'mmmm\\ d", "yyyy;@'
    ALIGNMENT = openpyxl.styles.Alignment(horizontal="left")
    """
    LSERM DOC
    | col  |    A   |   B   |  C    |
    |------+--------+-------+-------|
    | pt   | 365.45 | 98.55 | 48.15 |
    | pica |  30.37 |  8.21 |  4.01 |
    | code |  65.67 | 17.76 | 13.00 |
    """
    WIDTH_RIGHT = 17
    WIDTH_LEFT = WIDTH_RIGHT * 3
    wb: openpyxl.Workbook = None

    @classmethod
    def convert(cls, sup: OutputRecord, wb: openpyxl.Workbook):
        sup.__class__ = cls
        sup.wb = wb
        return sup

    def write(self):
        ws = self.wb.create_sheet(title=self.title())
        metadata_start = 1
        metadata_end = self.write_metadata(self._metadata, ws, metadata_start)

        table_start = metadata_end + 2
        table_1_col = 1
        table_2_col = table_1_col + 3

        data_start = self.write_header("Eligible ETFs (Recent Additions Highlighted)", ws,
                                       start_row=table_start, start_col=table_1_col)
        data_end = self.write_list(rows=self._etfs_eligible, with_matching=self._etfs_added, ws=ws,
                                   start_row=data_start, start_col=table_1_col)
        print(
            f"wrote table from ({table_start},{table_1_col})"
            f"to ({data_end},{table_1_col + 1}) in {ws.title}")

        data_start = self.write_header("Deletions", ws,
                                       start_row=table_start, start_col=table_2_col)
        data_end = self.write_list(rows=self._etfs_deleted, ws=ws,
                                   start_row=data_start, start_col=table_2_col)
        print(
            f"wrote table from ({table_start},{table_2_col})"
            f"to ({data_end},{table_2_col + 1}) in {ws.title}")

    @classmethod
    def write_metadata(cls, metadata: dict[str, str | datetime.date],
                       ws: openpyxl.worksheet.worksheet.Worksheet,
                       start_row: int):
        row = start_row
        for key in ["file_name", "quarter_end_date", "release_date", "effective_date"]:
            title = key.replace("_", " ").title()
            cell_left = ws.cell(row, 1, title)
            cell_right = ws.cell(row, 2, metadata[key])
            cell_left.font = cls.BOLD_FONT
            cell_right.font = cls.BOLD_FONT
            cell_right.number_format = cls.DATE_FORMAT
            cell_right.alignment = cls.ALIGNMENT
            row += 1
        return row

    @classmethod
    def write_header(cls, header: str, ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int,
                     start_col: int):
        header_title = ws.cell(start_row, start_col, header)

        header_left = ws.cell(start_row + 1, start_col, "Description")
        header_right = ws.cell(start_row + 1, start_col + 1, "SYMBOL")

        header_title.font = cls.BOLD_FONT
        header_left.font = cls.BOLD_FONT
        header_right.font = cls.BOLD_FONT

        header_left.border = cls.HEADER_LEFT
        header_right.border = cls.HEADER_RIGHT

        ws.column_dimensions[header_left.column_letter].width = cls.WIDTH_LEFT
        ws.column_dimensions[header_right.column_letter].width = cls.WIDTH_RIGHT

        return start_row + 2

    @classmethod
    def write_list(cls, rows: list[LsermRow], ws: openpyxl.worksheet.worksheet.Worksheet,
                   start_row: int, start_col: int,
                   with_matching: list[LsermRow] = ()):
        for idx, data in enumerate(rows):
            cell_a = ws.cell(start_row + idx, start_col)
            cell_b = cell_a.offset(0, 1)
            cell_a.value = data.col_a
            cell_b.value = data.col_b
            cell_a.border = cls.ROW_LEFT
            cell_b.border = cls.ROW_RIGHT
            for w in with_matching:
                if data.col_b == w.col_b:
                    cell_a.fill = OutputExcel.HIGHLIGHT
                    cell_b.fill = OutputExcel.HIGHLIGHT

        return start_row + len(rows)
