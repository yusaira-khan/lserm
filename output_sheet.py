import datetime

import openpyxl.worksheet.worksheet
import openpyxl.styles
import openpyxl.worksheet.dimensions

from lserm_row import LsermRow


class OutputSheet:
    BOLD_FONT = openpyxl.styles.Font("Calibri", bold=True)
    BORDER_SIDE = openpyxl.styles.Side(style="thin")
    HEADER_LEFT = openpyxl.styles.Border(top=BORDER_SIDE, bottom=BORDER_SIDE, left=BORDER_SIDE)
    HEADER_RIGHT = openpyxl.styles.Border(top=BORDER_SIDE, bottom=BORDER_SIDE, right=BORDER_SIDE)
    ROW_LEFT = openpyxl.styles.Border(left=BORDER_SIDE)
    ROW_RIGHT = openpyxl.styles.Border(right=BORDER_SIDE)
    HIGHLIGHT = openpyxl.styles.PatternFill("solid", bgColor="ffffffa6", fgColor="ffffffa6")
    WIDTH_RIGHT = 17
    WIDTH_LEFT = WIDTH_RIGHT * 3
    DATE_FORMAT = 'mmmm\\ d", "yyyy;@'
    ALIGNMENT = openpyxl.styles.Alignment(horizontal="left")

    def __init__(self, metadata=None):
        self._metadata = metadata
        self._all_current = []
        self._etfs_eligible = []
        self._all_eligible = []
        self._etfs_added = []
        self._all_deleted = []
        self._etfs_deleted = []

    def record_current(self, row: LsermRow):
        self.__record__(row, self._all_current, self._etfs_eligible)

    def record_addition(self, row: LsermRow):
        self.__record__(row, self._all_eligible, self._etfs_added)

    def record_deletion(self, row: LsermRow):
        self.__record__(row, self._all_deleted, self._etfs_deleted)

    def write(self, wb: openpyxl.Workbook):
        ws = wb.create_sheet(title=str(self._metadata["quarter_end_date"]))
        metadata_start = 1
        metadata_end = self.write_metadata(self._metadata, ws, metadata_start)

        table_start = metadata_end + 2
        table_1_col = 1
        table_2_col = table_1_col + 3

        data_start = self.write_header("Eligible ETFs (Recent Additions Highlighted)", ws,
                                       start_row=table_start, start_col=table_1_col)
        data_end = self.write_list(rows=self._etfs_eligible, with_matching=self._etfs_added, ws=ws,
                                   start_row=data_start, start_col=table_1_col)
        print(f"wrote table from ({table_start},{table_1_col}) to ({data_end},{table_1_col + 1}) in {ws.title}")

        data_start = self.write_header("Deletions", ws,
                                       start_row=table_start, start_col=table_2_col)
        data_end = self.write_list(rows=self._etfs_deleted, ws=ws,
                                   start_row=data_start, start_col=table_2_col)
        print(f"wrote table from ({table_start},{table_2_col}) to ({data_end},{table_2_col + 1}) in {ws.title}")

    @staticmethod
    def write_metadata(metadata: dict[str, str | datetime.date], ws: openpyxl.worksheet.worksheet.Worksheet,
                       start_row: int):
        row = start_row
        for key in ["file_name", "quarter_end_date", "release_date", "effective_date"]:
            title = key.replace("_", " ").title()
            cell_left = ws.cell(row, 1, title)
            cell_right = ws.cell(row, 2, metadata[key])
            cell_left.font = OutputSheet.BOLD_FONT
            cell_right.font = OutputSheet.BOLD_FONT
            cell_right.number_format = OutputSheet.DATE_FORMAT
            cell_right.alignment = OutputSheet.ALIGNMENT
            row += 1
        return row

    @staticmethod
    def write_header(header: str, ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int, start_col: int):
        header_title = ws.cell(start_row, start_col, header)

        header_left = ws.cell(start_row + 1, start_col, "Description")
        header_right = ws.cell(start_row + 1, start_col + 1, "SYMBOL")

        header_title.font = OutputSheet.BOLD_FONT
        header_left.font = OutputSheet.BOLD_FONT
        header_right.font = OutputSheet.BOLD_FONT

        header_left.border = OutputSheet.HEADER_LEFT
        header_right.border = OutputSheet.HEADER_RIGHT

        ws.column_dimensions[header_left.column_letter].width = OutputSheet.WIDTH_LEFT
        ws.column_dimensions[header_right.column_letter].width = OutputSheet.WIDTH_RIGHT

        return start_row + 2

    @staticmethod
    def write_list(rows: list[LsermRow], ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int, start_col: int,
                   with_matching: list[LsermRow] = ()):
        for idx, data in enumerate(rows):
            cell_a = ws.cell(start_row + idx, start_col)
            cell_b = cell_a.offset(0, 1)
            data.write(cell_a, cell_b)
            cell_a.border = OutputSheet.ROW_LEFT
            cell_b.border = OutputSheet.ROW_RIGHT
            for w in with_matching:
                if data.col_b == w.col_b:
                    cell_a.fill = OutputSheet.HIGHLIGHT
                    cell_b.fill = OutputSheet.HIGHLIGHT

        return start_row + len(rows)

    def print(self):
        print("Metadata", self._metadata)
        print("Current", len(self._all_current), len(self._etfs_eligible))
        self.print_table(self._all_current, self._etfs_eligible)
        print("Added", len(self._all_eligible), len(self._etfs_added))
        self.print_table(self._all_eligible, self._etfs_added)
        print("Deleted", len(self._all_deleted), len(self._etfs_deleted))
        self.print_table(self._all_deleted, self._etfs_deleted)

    @staticmethod
    def print_table(list1, list2):
        for i in list1:
            print(i, "(ETF)" if i in list2 else "")

    @staticmethod
    def __record__(row: LsermRow, all_list: list[LsermRow], etf_list: list[LsermRow]):
        all_list.append(row)
        if row.is_etf():
            etf_list.append(row)
